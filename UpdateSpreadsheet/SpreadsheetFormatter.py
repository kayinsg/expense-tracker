from openpyxl.styles import Alignment

from UpdateSpreadsheet.FontFormatter import FontFormatter
from UpdateSpreadsheet.dataObjects import FontProfile
from UpdateSpreadsheet.dataTypes import ExcelWorkbook, ExcelWorksheet


class SpreadsheetFormatter:
    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.dataRows = worksheet.max_row

    def apply(self):
        self.changeHeaderFont()
        self.changeBodyFont()
        self.adjustWidthOfColumnsToFit()
        self.padRowHeight()
        self.alignCellsBottomLeft()

    def changeHeaderFont(self) -> None:
        headerFontProfile = FontProfile(
            "Georgia",
            18,
            boldToggle = True
        )
        FontFormatter(self.worksheet).changeHeaderFont(headerFontProfile)

    def changeBodyFont(self) -> None:
        bodyFontProfile = FontProfile(
            "Helvetica Neue",
            12.8,
            boldToggle = False
        )
        FontFormatter(self.worksheet).changeBodyFont(bodyFontProfile)

    def adjustWidthOfColumnsToFit(self) -> None:
        for column in self.worksheet.iter_cols():
            currentColumnIndex = ""
            maxLengthOfCellInColumn = 0

            for cell in column:
                currentColumnIndex = cell.column_letter
                lengthOfCellValue = len(str(cell.value))
                if lengthOfCellValue > maxLengthOfCellInColumn:
                    maxLengthOfCellInColumn = lengthOfCellValue

            self.worksheet.column_dimensions[currentColumnIndex].width = maxLengthOfCellInColumn + 8

    def padRowHeight(self) -> None:
        idealRowHeight = 27
        for row in range(1, self.dataRows + 1):
            self.worksheet.row_dimensions[row].height = idealRowHeight

    def alignCellsBottomLeft(self) -> None:
        for row in self.worksheet.iter_rows(
                max_row=self.dataRows,
                max_col=self.worksheet.max_column
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="bottom"
                )
