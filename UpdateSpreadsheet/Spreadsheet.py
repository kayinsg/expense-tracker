from abc import abstractmethod
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet as ExcelWorksheet

from pandas import DataFrame
from pandas import Series
import pendulum


from UpdateSpreadsheet.FontFormatter import FontFormatter
from paths import spreadsheetPath
from abc import abstractmethod, ABC
from UpdateSpreadsheet.dataObjects import SpreadsheetDetails, FontProfile
from GlobalDataObjects import Data
from UpdateSpreadsheet.dataTypes import ExcelWorkbook, ExcelWorksheet


class Spreadsheet:
    def __init__(self, filePath: str):
        self.filePath = filePath

    def apply(self):
        formattedWorksheet = SpreadsheetFormatter(writtenWorksheet).apply()
        formattedWorksheet.workbook.save(self.filePath)
        return formattedWorksheet


class SpreadsheetInterface(ABC):
    @abstractmethod
    def apply(self) -> SpreadsheetDetails:
        raise NotImplementedError(
            'This is an abstract class.'
            'Desist from trying to instantiate'
        )


class SpreadsheetFormatter(SpreadsheetInterface):
    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.dataRows = worksheet.max_row

    def apply(self):
        self.changeHeaderFont()
        self.changeBodyFont()
        self.adjustWidthOfColumnsToFit()
        self.padRowHeight()
        self.alignCellsBottomLeft()

    def changeHeaderFont(self) -> None:
        headerFontProfile = FontProfile(
            "Georgia",
            18,
            boldToggle = True
        )
        FontFormatter(self.worksheet).changeHeaderFont(headerFontProfile)

    def changeBodyFont(self) -> None:
        bodyFontProfile = FontProfile(
            "Helvetica Neue",
            12.8,
            boldToggle = False
        )
        FontFormatter(self.worksheet).changeBodyFont(bodyFontProfile)

    def adjustWidthOfColumnsToFit(self) -> None:
        for column in self.worksheet.iter_cols():
            currentColumnIndex = ""
            maxLengthOfCellInColumn = 0

            for cell in column:
                currentColumnIndex = cell.column_letter
                lengthOfCellValue = len(str(cell.value))
                if lengthOfCellValue > maxLengthOfCellInColumn:
                    maxLengthOfCellInColumn = lengthOfCellValue

            self.worksheet.column_dimensions[currentColumnIndex].width = maxLengthOfCellInColumn + 8

    def padRowHeight(self) -> None:
        idealRowHeight = 27
        for row in range(1, self.dataRows + 1):
            self.worksheet.row_dimensions[row].height = idealRowHeight

    def alignCellsBottomLeft(self) -> None:
        lastDataColumnAccountedForRange = self.worksheet.max_column
        for row in self.worksheet.iter_rows(
                max_row=self.dataRows,
                max_col=lastDataColumnAccountedForRange
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="bottom"
                )
