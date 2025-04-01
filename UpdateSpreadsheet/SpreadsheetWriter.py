import openpyxl
import pandas

class WorksheetDataDepositor:
    def __init__(self, workbook: openpyxl.Workbook, worksheetName: str):
        self.workbook: openpyxl.Workbook = workbook
        self.worksheetName: str = worksheetName

    def insert(self, dataFrame: pandas.DataFrame, series: pandas.Series) -> openpyxl.Workbook:
        pandasDataRows: list[list] = self._getPandasDataRows(dataFrame, series)
        dateWorksheet = self._ensureWorksheetExistsInWorkbook(self.workbook, self.worksheetName)
        self._insertConsolidatedPandasDataInWorksheet(dateWorksheet, pandasDataRows)
        return self.workbook

    def _getPandasDataRows(self, dataFrame: pandas.DataFrame, series: pandas.Series) -> list[list]:
        return DataConsolidator(dataFrame).consolidate(series)

    def _ensureWorksheetExistsInWorkbook(self, workbook: openpyxl.Workbook, worksheetName: str) -> openpyxl.worksheet.worksheet.Worksheet:
        try:
            return workbook[worksheetName]
        except:
            return self.workbook.create_sheet(worksheetName)

    def _insertConsolidatedPandasDataInWorksheet(self, dateWorksheet, pandasDataRows: list[list]) -> None:
        for row in pandasDataRows:
            dateWorksheet.append(row)


class DataConsolidator:
    def __init__(self, dataFrame):
        self.dataFrame = dataFrame

    def consolidate(self, series):
        nestedListComponents = {
            'Data Frame Columns': self.getDataFrameColumns(self.dataFrame),
            'Data Frame Values': self.getDataFrameRowValues(self.dataFrame),
            'Series Headers': self.getSeriesHeaders(series),
        }
        nestedListComponents['Series Data'] = self.getSeriesRowValues(series, nestedListComponents['Series Headers'])        
        return self.normalizeNestedList(nestedListComponents)

    def getDataFrameColumns(self, dataFrame):
        return list(map(lambda x: x, dataFrame.columns.tolist()))

    def getDataFrameRowValues(self, dataFrame):
        return list(map(lambda dataStructure: dataStructure[1].tolist(),  dataFrame.iterrows()))

    def getSeriesHeaders(self, series):
        return list(map(lambda x: x, series.index.tolist()))

    def getSeriesRowValues(self, series, seriesHeaders):
        return list(map(lambda header: series.loc[header], seriesHeaders))

    def normalizeNestedList(self, nestedListComponents: dict) -> list[list]:
        return NestedListNormalizer(nestedListComponents).getFinalList()


class NestedListNormalizer:
    def __init__(self, nestedListComponents: dict):
        self.nestedListComponents = nestedListComponents
        self.finalNestedList: list[list] = [ ]

    def getFinalList(self):
        for key in self.nestedListComponents:
            rowValues = self.nestedListComponents[key]
            self._addRowValuesToFinalNestedList(rowValues)
        return self.finalNestedList

    def _addRowValuesToFinalNestedList(self, rowValues):
        typeOfList = self._typeOfListForPandasRowValues(rowValues)
        if typeOfList == 'Single List':
            self.finalNestedList.append(rowValues)
        elif typeOfList == 'List of Lists':
            self.finalNestedList.extend(rowValues)

    def _typeOfListForPandasRowValues(self, rowValues):
        if isinstance(rowValues[0], list):
            return 'List of Lists'
        return 'Single List'
