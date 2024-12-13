from abc import ABC, abstractmethod
import re as regex

from openpyxl.styles import Font
from .dataTransferObjects import RowInfo


class FontFormatter:
    def __init__(self, spreadsheetDetails):
        self.worksheet = spreadsheetDetails.worksheet
        self.workbook = spreadsheetDetails.workbook
        self.filePath = spreadsheetDetails.filePath

    def changeHeaderFont(self, FontProfile):
        HeaderFormatter(self.worksheet).changeFont(FontProfile)

    def changeBodyFont(self, FontProfile):
        BodyFormatter(self.worksheet).changeFont(FontProfile)

    def saveWorkbook(self):
        self.workbook.save(self.filePath)


class FontFormatterInterface(ABC):
    @abstractmethod
    def changeFont(self, FontProfile):
        raise NotImplementedError(
            'This is an abstract class.'
            'Desist from trying to instantiate'
        )


class HeaderFormatter:
    def __init__(
        self,
        worksheet,
    ):
        self.worksheet = worksheet
        self.headerRowNumbers = RowIdentifier(worksheet).fetchHeaderRowNumbers()

    def changeFont(self, FontProfile):

        worksheet = self.worksheet
        lastColumnContainingData = worksheet.max_column + 1
        for headerRow in self.headerRowNumbers:
            for columnNumber in range(1, lastColumnContainingData):

                headerCell = worksheet.cell(row=headerRow, column=columnNumber)
                headerCell.font = Font(
                    name=FontProfile.name,
                    size=FontProfile.size,
                    bold=FontProfile.boldToggle
                )


class BodyFormatter:
    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.bodyRowNumbers = RowIdentifier(
            worksheet
        ).fetchBodyRowNumbers()

    def changeFont(self, FontProfile):
        worksheet = self.worksheet
        lastColumnContainingData = worksheet.max_column + 1
        for bodyRow in self.bodyRowNumbers:
            for columnNumber in range(
                1,
                lastColumnContainingData
            ):
                headerCell = worksheet.cell(
                    row=bodyRow,
                    column=columnNumber
                )
                headerCell.font = Font(
                    name=FontProfile.name,
                    size=FontProfile.size,
                    bold=FontProfile.boldToggle
                )


class RowIdentifier:
    def __init__(self, worksheet):
        self.worksheet = worksheet

    def fetchHeaderRowNumbers(self) -> list[int]:
        return self.classifyRows().headerRowNumbers

    def fetchBodyRowNumbers(self) -> list[int]:
        return self.classifyRows().bodyRowNumbers

    def classifyRows(self):
        rowDetails = RowInfo([], [])
        for row in self.worksheet.iter_rows():
            cellValues = list()

            for cell in row:
                cellValues.append(cell.value)
            if self.rowComprisesHeader(cellValues):
                for cell in row:
                    rowDetails.headerRowNumbers.append(cell.row)
                    break
            else:
                for cell in row:
                    rowDetails.bodyRowNumbers.append(cell.row)
                    break

        return rowDetails

    def rowComprisesHeader(self, rowValues):

        headerPattern = regex.compile(r'\w')
        checkedValues = list()

        for element in rowValues:
            if headerPattern.match(element):
                checkedValues.append(True)
            else:
                checkedValues.append(False)

        if all(checkedValues):
            return True

        return False
