from abc import ABC, abstractmethod
from openpyxl.styles import Font


class FontFormatter:
    def __init__(self,worksheet):
        self.worksheet = worksheet

    def changeHeaderFont(self, FontProfile) -> None:
        headerRowNumbers = self.getRowNumbers('header')
        HeaderFormatter(self.worksheet, headerRowNumbers).changeFont(FontProfile)

    def changeBodyFont(self, FontProfile) -> None:
        headerRowNumbers = self.getRowNumbers('body')
        BodyFormatter(self.worksheet, headerRowNumbers).changeFont(FontProfile)

    def getRowNumbers(self, rowType):
        return TypeOfRowIdentifier(self.worksheet).fetchRowNumbers(rowType)


class FontFormatterInterface(ABC):
    @abstractmethod
    def changeFont(self, FontProfile) -> None:
        raise NotImplementedError(
            'This is an abstract class.'
            'Desist from trying to instantiate'
        )

    def getColumnNumbersThatContainData(self, worksheet):
        lastNonEmptyColumnAccountedForRange = worksheet.max_column + 1
        return list(map(lambda columnNumber: columnNumber, range(1, lastNonEmptyColumnAccountedForRange)))


class HeaderFormatter(FontFormatterInterface):
    def __init__(self, worksheet, headerRowNumbers):
        self.worksheet = worksheet
        self.headerRowNumbers = headerRowNumbers

    def changeFont(self, FontProfile) -> None:
        for headerRow in self.headerRowNumbers:
            for columnNumber in self.getColumnNumbersThatContainData(self.worksheet):
                cellCoordinates = {'rowNumber':headerRow, 'columnNumber': columnNumber}
                self.changeHeaderSpreadsheetCells(FontProfile, cellCoordinates)

    def changeHeaderSpreadsheetCells(self, fontProfile, cellCoordinates):
        CellFontChanger(self.worksheet, fontProfile).change(cellCoordinates)


class BodyFormatter(FontFormatterInterface):
    def __init__(self, worksheet, bodyRowNumbers):
        self.worksheet = worksheet
        self.bodyRowNumbers = bodyRowNumbers

    def changeFont(self, FontProfile) -> None:
        for bodyRow in self.bodyRowNumbers:
            for columnNumber in self.getColumnNumbersThatContainData(self.worksheet):
                cellCoordinates = {'rowNumber':bodyRow, 'columnNumber': columnNumber}
                self.changeBodySpreadsheetCells(FontProfile, cellCoordinates)

    def changeBodySpreadsheetCells(self, fontProfile, cellCoordinates):
        CellFontChanger(self.worksheet, fontProfile).change(cellCoordinates)


class CellFontChanger:
    def __init__(self, worksheet, fontProfile):
        self.worksheet = worksheet
        self.fontProfile = fontProfile

    def change(self, cellCoordinates):
        cellObject = self.getCellInstance(cellCoordinates)
        self.changeCellFontProperties(self.fontProfile, cellObject)

    def getCellInstance(self, cellCoordinates):
        bodyCell = self.worksheet.cell(
            row=cellCoordinates['rowNumber'],
            column=cellCoordinates['columnNumber']
        )
        return bodyCell

    def changeCellFontProperties(self, FontProfile, bodyCell):
        bodyCell.font = Font(
            name=FontProfile.name,
            size=FontProfile.size,
            bold=FontProfile.boldToggle
        )


class TypeOfRowIdentifier:
    def __init__(self, worksheet):
        self.worksheet = worksheet

    def fetchRowNumbers(self, rowType):
        worksheetData = self.getRowsInWorksheet(self.worksheet)
        if rowType == "header":
            return self.getHeaderRowNumbers(worksheetData)
        elif rowType == "body":
            return self.getBodyRowNumbers(worksheetData)

    def getRowsInWorksheet(self, worksheet):
        return list(map(
            lambda rowInSpreadsheet: rowInSpreadsheet,
            worksheet.iter_rows(values_only=True)
        ))

    def getHeaderRowNumbers(self, worksheetDataRows):
        headerRowNumbers = [ ]
        for index, row in enumerate(worksheetDataRows, start=1):
            if self.typeOfRow(row) == "Header":
                headerRowNumbers.append(index)
        return headerRowNumbers

    def getBodyRowNumbers(self, worksheetDataRows):
        bodyRowNumbers = [ ]
        for index, row in enumerate(worksheetDataRows, start=1):
            if self.typeOfRow(row) == "Body":
                bodyRowNumbers.append(index)
        return bodyRowNumbers

    def typeOfRow(self, row):
        for item in row:
            if '$' in str(item):
                return 'Body'
        return 'Header'
