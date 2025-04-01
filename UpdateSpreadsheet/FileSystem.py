from abc import ABC, abstractmethod
import pendulum
import os
import openpyxl


class FileSystem:
    def __init__(self, spreadsheetParentDirectory):
        self.spreadsheetParentDirectory = spreadsheetParentDirectory

    def setUpSpreadsheet(self, date):
        monthFolder = self.createSpreadsheetMonthFolder(date)
        workbook = self.createSpreadsheetFile(monthFolder, date)
        return workbook

    def createSpreadsheetMonthFolder(self, date):
        directoryCreator = DirectoryCreator(self.spreadsheetParentDirectory)
        createdMonthDirectory = MonthDirectory(directoryCreator, date).create()
        return createdMonthDirectory

    def createSpreadsheetFile(self, parentDirectory, date):
        fileCreator = FileCreator(parentDirectory)
        return SpreadsheetFileCreator(fileCreator, date).create()


class FileSystemInterface(ABC):

    @staticmethod
    def standardizePath(path):
        if not path.endswith("/"):
            path += "/"
        return path

    @abstractmethod
    def create(self):
        pass


class MonthDirectory(FileSystemInterface):
    def __init__(self, folderCreator, currentDate):
        self.folderCreator = folderCreator
        self.currentDate = currentDate

    def create(self):
        monthOfCurrentDate = self.getMonthOfCurrentDate(self.currentDate)
        return self.makeFolderRepresentingTheMonthOfSpreadsheet(monthOfCurrentDate)

    def getMonthOfCurrentDate(self, currentDate):
        return pendulum.parse(currentDate).format('MMMM')

    def makeFolderRepresentingTheMonthOfSpreadsheet(self, monthName):
        spreadsheetMonthDirectory = self.folderCreator.createDirectory(monthName)
        return self.standardizePath(spreadsheetMonthDirectory)


class DirectoryCreator:
    def __init__(self, parentDirectory):
        self.parentDirectory = parentDirectory

    def createDirectory(self, monthName):
        fullFilePath = self.createFullPathForDirectory(monthName)
        self.makeDirectoryOnFileSystem(fullFilePath)
        return fullFilePath 

    def createFullPathForDirectory(self, monthName):
        return f"{self.parentDirectory}{monthName}"

    def makeDirectoryOnFileSystem(self, monthName):
        try:
            os.makedirs(monthName)
            print(f"Folder '{monthName}' created successfully.")
        except FileExistsError:
            print(f"Folder '{monthName}' already exists.")
        except Exception as e:
            print(f"An error occurred while creating the folder: {e}")


class SpreadsheetFileCreator(FileSystemInterface):
    def __init__(self, fileCreator, currentDate):
        self.fileCreator = fileCreator
        self.currentDate = currentDate

    def create(self):
        weekWithinMonth = self.getWeekInMonth(self.currentDate)
        spreadsheetFileName = self.getSpreadsheetFileName(weekWithinMonth)
        spreadsheetFile =  self.createSpreadsheetOnFileSystem(spreadsheetFileName)
        workbook = self.getWorkbook(spreadsheetFile)
        return {'FilePath': spreadsheetFile, 'Workbook': workbook}

    def getWeekInMonth(self, isoDate):
        date = pendulum.parse(isoDate)
        return date.week_of_month

    def getSpreadsheetFileName(self, weekWithinMonth):
        return f'Week {weekWithinMonth}'

    def createSpreadsheetOnFileSystem(self, spreadsheetFileName):
        return self.fileCreator.createFile(spreadsheetFileName)

    def getWorkbook(self, spreadsheetFilePath):
        return openpyxl.load_workbook(spreadsheetFilePath)


class FileCreator:
    def __init__(self, parentDirectory):
        self.parentDirectory = parentDirectory

    def createFile(self, fileName):
       completeFileName = self.completeFileName(fileName) 
       self.createSpreadsheetFile(completeFileName)
       return completeFileName

    def completeFileName(self, fileName):
        return f"{self.parentDirectory}{fileName}.xlsx"

    def createSpreadsheetFile(self, completeFileName):
        workbook = openpyxl.Workbook()
        workbook.save(completeFileName)
