import os
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet as ExcelWorksheet

from pandas import DataFrame
from pandas import Series
import pendulum


from SupportInterfaces.FontFormatter import FontFormatter
from SupportInterfaces.TableConstructor import TableCreator
from SupportInterfaces.TypeChecker import TypeChecker
from SupportInterfaces.SummaryConstructor import Summary
from SupportInterfaces.dataTransferObjects import FontProfile, SpreadsheetDetails
from SupportInterfaces.utils import readLinesFromFile
from paths import flatTextFile, spreadsheetPath


class TableFacade:
    def __init__(self, flatTextPath: str):
        self.itemPricePairs = DataExtractor(
            readLinesFromFile(flatTextPath)
        ).categorizeData()

    def getRawTable(self) -> DataFrame:
        return TableCreator(self.itemPricePairs).makeTable("raw")

    def getFormattedTable(self) -> DataFrame:
        return TableCreator(self.itemPricePairs).makeTable("view")


class DataExtractor:
    def __init__(self, listOfLinesInFile: list[str]):
        self.listOfLinesInFile = listOfLinesInFile

    def categorizeData(self) ->  list[tuple[str, int | float, float, int]]:
        listOfLinesInFile = self.listOfLinesInFile

        items: list[str] = list(filter(
            self._lineComprisesStrings,
            listOfLinesInFile
        ))
        pricesRepresentedAsStrings = list(filter(
            self._lineComprisesNumbers,
            listOfLinesInFile
        ))
        numericPrices = list(map(
            self._convertPricesToNumericDataType,
            pricesRepresentedAsStrings
        ))
        afterTaxPrices = list(map(
            self._calculateAfterTaxPrices,
            numericPrices
        ))

        taxPaidPerItem = self._calculateTaxesPaidPerItem(
            numericPrices,
            afterTaxPrices
        )

        itemPriceInformation = list(zip(
            items,
            numericPrices,
            afterTaxPrices,
            taxPaidPerItem
        ))

        return itemPriceInformation

    def _lineComprisesStrings(self, entry: str) -> bool:
        entryType = TypeChecker(entry).dataType
        if entryType == "String":
            return True
        return False

    def _lineComprisesNumbers(self, entry: str) -> bool:
        entryType = TypeChecker(entry).dataType
        if entryType == "Integer":
            return True
        elif entryType == "Decimal":
            return True
        return False

    def _convertPricesToNumericDataType(self, price: str) -> int | float:
        dataTypeOfPrice = TypeChecker(price).dataType
        if dataTypeOfPrice == "Integer":
            return int(price)
        elif dataTypeOfPrice == "Decimal":
            return float(price)
        return 0

    def _calculateAfterTaxPrices(self, price: int | float):
        countryTaxRateInDecimal = 0.13
        taxMultiplier = 1 + countryTaxRateInDecimal
        afterTaxPrice = taxMultiplier * price
        return afterTaxPrice

    def _calculateTaxesPaidPerItem(
        self,
        grossPrices: list[int | float],
        afterTaxPrices: list[int | float],
    ) -> list[int] :

        pricePairs = list(zip(grossPrices, afterTaxPrices))

        taxesPaidPerItem = list()

        for grossPrice, afterTaxPrice in pricePairs:
            taxPaid = afterTaxPrice - grossPrice
            taxesPaidPerItem.append(taxPaid)

        return taxesPaidPerItem


class SpreadsheetCreator:

    @classmethod
    def workbook(cls, filename: str) -> openpyxl.Workbook:
        try:
            return load_workbook(
                filename, keep_vba=True, keep_links=True
            )
        except ValueError as InvalidFileException:
            print(InvalidFileException)
            print('[ ERROR ] File Not Found Creating One Now')
            return openpyxl.Workbook()

    def get(self, filePath: str) -> SpreadsheetDetails:
        workbook = SpreadsheetCreator.workbook(filePath)
        self._removeUndesiredWorksheets(workbook)
        worksheet = self._createDateWorksheet(workbook)
        workbook.save(spreadsheetPath)
        return SpreadsheetDetails(
            filePath,
            workbook,
            worksheet
        )

    def _removeUndesiredWorksheets(self, workbook):
        worksheetKeyword = "Budget"
        listOfWorksheets = workbook.worksheets
        for worksheet in listOfWorksheets:
            worksheetName = str(worksheet)
            if worksheetKeyword in worksheetName:
                pass
            else:
                workbook.remove(worksheet)

    def _createDateWorksheet(self, workbook: openpyxl.Workbook) -> ExcelWorksheet:
        formattedCurrentDate = pendulum.now().format("MMM.DD.YYYY")
        worksheetName = f"{formattedCurrentDate} Budget"
        dateWorksheet = workbook.create_sheet(worksheetName, 0)
        return dateWorksheet


class SpreadsheetWriter:
    def __init__(
        self,
        spreadsheetDetails: SpreadsheetDetails,
        table: DataFrame,
        summary: Series
    ):
        self.spreadsheetDetails = spreadsheetDetails
        self.worksheet = spreadsheetDetails.worksheet
        self.table = table
        self.summary = summary

    def captureStateAfterWrite(self) -> SpreadsheetDetails:
        table = self.table
        summary = self.summary

        workbook = self.spreadsheetDetails.workbook
        spreadSheetFile = self.spreadsheetDetails.filePath

        self._writeTabularDataToWorksheet(table)
        self.writeSummaryToWorksheet(summary)

        workbook.save(spreadSheetFile)

        return SpreadsheetDetails(
            spreadSheetFile,
            workbook,
            self.spreadsheetDetails.worksheet
        )

    def _writeTabularDataToWorksheet(self, table: DataFrame) -> None:
        worksheet = self.spreadsheetDetails.worksheet
        workbook = self.spreadsheetDetails.workbook

        workbook.active = worksheet
        for row in dataframe_to_rows(
            table,
            index=False,
            header=True
        ):

            worksheet.append(row)

    def writeSummaryToWorksheet(self, series: Series) -> None:
        summaryHeaders = list()
        summaryValues = list()

        indices = series.index.tolist()

        for index in indices:
            summaryHeaders.append(index)
            valueForSummary = series[index]
            summaryValues.append(valueForSummary)

        self.spreadsheetDetails.worksheet.append(summaryHeaders)
        self.spreadsheetDetails.worksheet.append(summaryValues)


class SpreadsheetFormatter:
    def __init__(
        self,
        spreadsheetDetails : SpreadsheetDetails,
    ):

        self.spreadsheetDetails = spreadsheetDetails
        self.dataRows = self.spreadsheetDetails.worksheet.max_row

    def fetchFormattedWorkSheet(self) -> SpreadsheetDetails:
        return self.formatColumns()

    def formatColumns(self) -> SpreadsheetDetails:
        workbook = self.spreadsheetDetails.workbook
        filePath = self.spreadsheetDetails.filePath

        self.changeHeaderFont()
        self.changeBodyFont()
        self.adjustWidthOfColumnsToFit()
        self.padRowHeight()
        self.alignCellsBottomLeft()
        workbook.save(filePath)

        return SpreadsheetDetails(
            filePath,
            workbook,
            self.spreadsheetDetails.worksheet,
        )

    def changeHeaderFont(self) -> None:
        headerFontProfile = FontProfile(
            "Georgia",
            18,
            boldToggle = True
        )
        fontFormatter = FontFormatter(self.spreadsheetDetails)
        fontFormatter.changeHeaderFont(headerFontProfile)

    def changeBodyFont(self) -> None:
        bodyFontProfile = FontProfile(
            "Helvetica Neue",
            12.8,
            boldToggle = False
        )
        fontFormatter = FontFormatter(self.spreadsheetDetails)
        fontFormatter.changeBodyFont(bodyFontProfile)

    def adjustWidthOfColumnsToFit(self) -> None:
        worksheet = self.spreadsheetDetails.worksheet

        for column in worksheet.iter_cols():
            currentColumnIndex = ""
            maxLengthOfCellInColumn = 0

            for cell in column:
                currentColumnIndex = cell.column_letter
                lengthOfCellValue = len(str(cell.value))
                if lengthOfCellValue > maxLengthOfCellInColumn:
                    maxLengthOfCellInColumn = lengthOfCellValue

            worksheet.column_dimensions[currentColumnIndex].width = maxLengthOfCellInColumn + 8

    def padRowHeight(self) -> None:
        worksheet = self.spreadsheetDetails.worksheet
        idealRowHeight = 27

        dataRows = worksheet.max_row + 1

        for row in range(1, dataRows):
            worksheet.row_dimensions[row].height = idealRowHeight

    def alignCellsBottomLeft(self) -> None:
        worksheet = self.spreadsheetDetails.worksheet
        lastDataColumnAccountedForRange = worksheet.max_column

        for row in worksheet.iter_rows(
                max_row=self.dataRows,
                max_col=lastDataColumnAccountedForRange
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="bottom"
                )


def getFormattedData():
    flatTextFileContent = readLinesFromFile(flatTextFile)
    itemPricePairs = DataExtractor(flatTextFileContent).categorizeData()
    summary = Summary(itemPricePairs)
    formattedSummary = summary.getFormattedSummary()
    
    table = TableFacade(flatTextFile)
    viewTable = table.getFormattedTable()
    
    return {
        "Formatted Table": viewTable,
        "Formatted Summary": formattedSummary,
    }

def getRawData():
    rawTable = TableFacade(flatTextFile).getRawTable()
    itemPricePairs = DataExtractor(readLinesFromFile(flatTextFile)).categorizeData()
    summary = Summary(itemPricePairs)
    rawSummary = summary.getRawSummary()
    
    return {
        "Raw Table": rawTable,
        "Raw Summary": rawSummary,
        }


def createDateWorksheet():
    return SpreadsheetCreator().get(spreadsheetPath)


def writtenSpreadSheet(spreadsheetDetails):
    formattedData = getFormattedData()
    formattedTable = formattedData['Formatted Table']
    formattedSummary = formattedData['Formatted Summary']

    spreadsheetToBeWrittenTo = SpreadsheetWriter(
        spreadsheetDetails,
        formattedTable,
        formattedSummary

    )
    writtenSpreadSheet = spreadsheetToBeWrittenTo.captureStateAfterWrite()
    return writtenSpreadSheet


def formattedSpreadSheet(spreadsheetDetails):
    spreadsheetToBeFormatted = SpreadsheetFormatter(spreadsheetDetails)
    formattedSpreadsheetDetails = spreadsheetToBeFormatted.fetchFormattedWorkSheet()
    return formattedSpreadsheetDetails.workbook.save(
        spreadsheetDetails.filePath
    )


formattedSpreadSheet(writtenSpreadSheet(createDateWorksheet()))
