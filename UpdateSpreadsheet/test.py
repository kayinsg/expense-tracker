import unittest
from openpyxl import Workbook
from FileSystem import MonthDirectory, DirectoryCreator, SpreadsheetFileCreator, FileCreator
from WorkbookPopulator import WorkbookPopulator
import pendulum
from colour_runner.runner import ColourTextTestRunner
import pandas
from WorksheetWriter import WorksheetDataDepositor
from DateTranslator import getCurrentDateObject


class SpreadsheetFileSystemTest(unittest.TestCase):

    currentDate = '2025-03-20'

    def testShouldCreateFolderRepresentingMonthWhenDateIsGiven(self):

        class FakeDirectoryCreator(DirectoryCreator):
            def __init__(self, parentDirectory):
                self.parentDirectory = parentDirectory

            def makeDirectoryOnFileSystem(self, monthName):
                return monthName

        parentDirectory = ""
        directoryCreator = FakeDirectoryCreator(parentDirectory)
        monthDirectory = MonthDirectory(directoryCreator, self.currentDate)

        fullPath = monthDirectory.create()

        self.assertEqual(fullPath, "March/")

    def testShouldCreateSpreadsheetFileRepresentingTheWeekNumberInMonthGivenCurrentDate(self):

        class FakeSpreadsheetFileCreator(SpreadsheetFileCreator):
            def __init__(self, fileCreator, currentDate):
                self.fileCreator = fileCreator
                self.currentDate = currentDate

            def getWorkbook(self, spreadsheetFilePath):
                return spreadsheetFilePath

        class FakeFileCreator(FileCreator):
            def __init__(self, parentDirectory):
                self.parentDirectory = parentDirectory

            def createSpreadsheetFile(self, filePath):
                return filePath

        def get_week_in_month(iso_date: str) -> int:
            date = pendulum.parse(iso_date)
            return date.week_of_month

        parentDirectory = ""
        fileCreator = FakeFileCreator(parentDirectory)
        weekWithinMonth = get_week_in_month(self.currentDate)

        spreadsheet = FakeSpreadsheetFileCreator(fileCreator, self.currentDate).create()

        self.assertEqual(spreadsheet['FilePath'], f"{parentDirectory}Week {weekWithinMonth}.xlsx")


class WorksheetPopulationTest(unittest.TestCase):

        @staticmethod
        def getWorkbook():
            return Workbook()

        def testShouldAccountForDaysOfWeekGivenDaysFromPreviousMonth(self):
            dateWithPreviousMonthWeekdays = '2025-03-01'
            weekDaysInvolvingThePreviousMonth = ['Feb.23.2025', 'Feb.24.2025', 'Feb.25.2025', 'Feb.26.2025', 'Feb.27.2025', 'Feb.28.2025', 'Mar.01.2025']
            workbook = self.getWorkbook()

            workbookWithDateWorksheets = WorkbookPopulator(dateWithPreviousMonthWeekdays).populate(workbook)

            self.assertEqual(workbookWithDateWorksheets.sheetnames, weekDaysInvolvingThePreviousMonth)

        def testShouldAccountForDaysOfWeekGivenDaysFromTheSameMonth(self):
            dateWithSameMonthWeekdays = '2025-03-27'
            weekDaysInvolvingOnlyTheSameMonth = ['Mar.23.2025', 'Mar.24.2025', 'Mar.25.2025', 'Mar.26.2025', 'Mar.27.2025', 'Mar.28.2025', 'Mar.29.2025']
            workbook = self.getWorkbook()

            workbookWithDateWorksheets = WorkbookPopulator(dateWithSameMonthWeekdays).populate(workbook)

            self.assertEqual(workbookWithDateWorksheets.sheetnames, weekDaysInvolvingOnlyTheSameMonth)

        def testShouldAccountForDaysOfWeekGivenDaysFromSucceedingMonth(self):
            dateWithSuceedingMonthWeekdays = '2025-04-05'
            weekDaysInvolvingTheSucceedingMonth = ['Mar.30.2025', 'Mar.31.2025', 'Apr.01.2025', 'Apr.02.2025', 'Apr.03.2025', 'Apr.04.2025', 'Apr.05.2025']
            workbook = self.getWorkbook()

            workbookWithDateWorksheets = WorkbookPopulator(dateWithSuceedingMonthWeekdays).populate(workbook)

            self.assertEqual(workbookWithDateWorksheets.sheetnames, weekDaysInvolvingTheSucceedingMonth)


class InsertionOfDataInSpreadsheetTest(unittest.TestCase):

    def testShouldPlaceDataframeAndSeriesInSpreadsheetAfterBeingConvertedToNestedList(self):

        def getSheetData(workbook, worksheetName):
            if worksheetName not in workbook.sheetnames:
                raise ValueError(f"Worksheet '{worksheetName}' not found in workbook")
            worksheet = workbook[worksheetName]
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(list(row))
            return data

        def getNestedList():
            return [
                ['Items', 'Gross Price', 'Final Price', 'Taxes Paid'],
                ['Candy', 1.00, 1.13, 0.13],
                ['Orange', 1.50, 1.76, 0.25],
                ['Chips', 1.99, 2.13, 0.35],
                ['Total Items', 'Total Gross Price', 'Total Final Price', 'Total Taxes Paid'],
                [3.00, 4.49, 5.02, 0.73]
            ]

        workbook = Workbook()
        worksheetName = 'Sheet 1'
        nestedList = getNestedList()

        workbookWithData = WorksheetDataDepositor(workbook, worksheetName).insert(nestedList)
        dataInWorksheet = getSheetData(workbookWithData, worksheetName)

        self.assertEqual(dataInWorksheet, nestedList)


class DateConversionTests(unittest.TestCase):

    def getFakeDateFromPendulum(self):
        return '2025-03-23'

    def testShouldConvertISODateToSpreadsheetDate(self):
        # GIVEN the following preconditions corresponding to the system under test:
        expectedResult = 'Mar.23.2025'
        # WHEN the following module is executed:
        actualResult = getCurrentDateObject(self.getFakeDateFromPendulum)
        # THEN the observable behavior should be verified as stated below:
        self.assertEqual(actualResult.spreadsheet, expectedResult)

    def testShouldConvertSpreadsheetDateToISO(self):
        # GIVEN the following preconditions corresponding to the system under test:
        expectedResult = '2025-03-23'
        # WHEN the following module is executed:
        actualResult = getCurrentDateObject(self.getFakeDateFromPendulum)
        # THEN the observable behavior should be verified as stated below:
        self.assertEqual(actualResult.iso, expectedResult)

if __name__ == '__main__':
    unittest.main(testRunner=ColourTextTestRunner(), verbosity=2)
