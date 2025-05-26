import openpyxl

class WorksheetDataDepositor:
    def __init__(self, workbook: openpyxl.Workbook, worksheetName: str):
        self.workbook: openpyxl.Workbook = workbook
        self.worksheetName: str = worksheetName

    def insert(self, budgetInfo) -> openpyxl.Workbook:
        dateWorksheet = self._ensureWorksheetExistsInWorkbook(self.workbook, self.worksheetName)
        self._insertConsolidatedPandasDataInWorksheet(dateWorksheet, budgetInfo)
        return self.workbook

    def _ensureWorksheetExistsInWorkbook(self, workbook: openpyxl.Workbook, worksheetName: str) -> openpyxl.worksheet.worksheet.Worksheet:
        try:
            return workbook[worksheetName]
        except:
            return self.workbook.create_sheet(worksheetName)

    def _insertConsolidatedPandasDataInWorksheet(self, dateWorksheet, budgetInfo: list[list]) -> None:
        for row in budgetInfo:
            dateWorksheet.append(row)
